import logging
import shutil
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from automations.drive.spreadsheet_schema import REQUIRED_COLUMNS
from config import (
    MAX_SPREADSHEET_CELLS,
    MAX_SPREADSHEET_COLUMNS,
    MAX_SPREADSHEET_ROWS,
    MAX_UPLOAD_SIZE_BYTES,
    MAX_XLSX_UNCOMPRESSED_SIZE_BYTES,
    UPLOAD_TEMP_DIR,
)

logger = logging.getLogger(__name__)
ALLOWED_EXTENSION = ".xlsx"


class UploadValidationError(ValueError):
    def __init__(self, message: str, status_code: int = 400):
        super().__init__(message)
        self.status_code = status_code


@dataclass(frozen=True)
class PreparedUpload:
    input_file: Path
    temporary_dir: Path
    original_filename: str
    size_bytes: int


def prepare_xlsx_upload(file_storage: FileStorage | None) -> PreparedUpload:
    if file_storage is None or not file_storage.filename:
        raise ValueError(
            "Selecione um arquivo .xlsx para executar a automação.",
            400,
        )

    original_filename = Path(file_storage.filename).name
    suffixes = Path(original_filename).suffixes
    if len(suffixes) != 1 or suffixes[0].casefold() != ALLOWED_EXTENSION:
        raise ValueError("Formato inválido. Envie somente arquivos .xlsx.", 400)

    execution_dir = UPLOAD_TEMP_DIR / uuid4().hex
    execution_dir.mkdir(parents=True, exist_ok=False)
    input_file = execution_dir / "input.xlsx"

    try:
        file_storage.save(input_file)
        size_bytes = input_file.stat().st_size
        if size_bytes <= 0:
            raise UploadValidationError(
                "Selecione um arquivo .xlsx para executar a automação.",
                400,
            )
        if size_bytes > MAX_UPLOAD_SIZE_BYTES:
            raise UploadValidationError(
                "Arquivo acima do limite permitido.",
                413,
            )

        validate_xlsx_file(input_file)
        return PreparedUpload(
            input_file=input_file,
            temporary_dir=execution_dir,
            original_filename=original_filename,
            size_bytes=size_bytes,
        )
    except UploadValidationError:
        cleanup_upload_dir(execution_dir)
        raise
    except Exception as exc:
        cleanup_upload_dir(execution_dir)
        raise UploadValidationError(
            "O arquivo enviado não é uma planilha .xlsx válida.",
            422,
        ) from exc


def validate_xlsx_file(input_file: Path) -> None:
    validate_xlsx_archive(input_file)
    workbook = None
    try:
        workbook = load_workbook(
            filename=input_file,
            read_only=True,
            data_only=True,
        )
        worksheet = workbook.active
        header_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                max_col=MAX_SPREADSHEET_COLUMNS + 1,
                values_only=True,
            ),
            None,
        )
        if (
            header_row
            and len(header_row) > MAX_SPREADSHEET_COLUMNS
            and any(header_row[MAX_SPREADSHEET_COLUMNS:])
        ):
            raise UploadValidationError(
                "A planilha excede o limite de colunas permitido.", 422
            )

        headers = {
            str(value).strip() for value in header_row or [] if value is not None
        }
        missing_columns = [
            column for column in REQUIRED_COLUMNS if column not in headers
        ]
        if missing_columns:
            missing_label = ", ".join(missing_columns)
            raise UploadValidationError(
                f"A planilha não contém as colunas obrigatórias: {missing_label}.",
                422,
            )

        servico_index = next(
            index
            for index, value in enumerate(header_row)
            if value is not None and str(value).strip() == "SERVICO"
        )
        processed_rows = 0
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=MAX_SPREADSHEET_ROWS + 2,
            max_col=MAX_SPREADSHEET_COLUMNS,
            values_only=True,
        ):
            if len(row) <= servico_index or row[servico_index] is None:
                break
            processed_rows += 1

            if processed_rows * MAX_SPREADSHEET_COLUMNS > MAX_SPREADSHEET_CELLS:
                raise UploadValidationError(
                    "A planilha excede o limite de células permitido.", 422
                )
            if processed_rows > MAX_SPREADSHEET_ROWS:
                raise UploadValidationError(
                    "A planilha excede o limite de linhas permitido.", 422
                )

        if processed_rows == 0:
            raise UploadValidationError(
                "A planilha não possui registros para processamento.", 422
            )
    finally:
        if workbook is not None:
            workbook.close()


def validate_xlsx_archive(input_file: Path) -> None:
    if not zipfile.is_zipfile(input_file):
        raise UploadValidationError(
            "O arquivo enviado não é uma planilha .xlsx válida.", 422
        )

    total_uncompressed_size = 0
    try:
        with zipfile.ZipFile(input_file) as archive:
            for info in archive.infolist():
                total_uncompressed_size += info.file_size
                if total_uncompressed_size > MAX_XLSX_UNCOMPRESSED_SIZE_BYTES:
                    raise UploadValidationError(
                        "A planilha excede o limite de tamanho permitido.", 422
                    )
    except zipfile.BadZipFile as exc:
        raise UploadValidationError(
            "O arquivo enviado não é uma planilha .xlsx válida.",
            422,
        ) from exc


def cleanup_upload(upload: PreparedUpload | None) -> None:
    if upload is None:
        return
    cleanup_upload_dir(upload.temporary_dir)


def cleanup_upload_dir(directory: Path) -> None:
    try:
        resolved_upload_root = UPLOAD_TEMP_DIR.resolve()
        resolved_directory = directory.resolve()
        if resolved_directory == resolved_upload_root:
            return
        if resolved_upload_root not in resolved_directory.parents:
            logger.warning("Ignorando limpeza fora do diretório de uploads controlado")
            return
        shutil.rmtree(resolved_directory, ignore_errors=False)
    except FileNotFoundError:
        return
    except Exception:
        logger.exception("Falha ao remover diretório temporário de upload")


def cleanup_stale_uploads(max_age_seconds: int = 24 * 60 * 60) -> None:
    UPLOAD_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    cutoff = time.time() - max_age_seconds

    for child in UPLOAD_TEMP_DIR.iterdir():
        try:
            if child.is_dir() and child.stat().st_mtime < cutoff:
                cleanup_upload_dir(child)
        except FileNotFoundError:
            continue
        except Exception:
            logger.exception("Falha ao limpar upload temporário antigo")
